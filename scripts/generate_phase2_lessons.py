"""Generate the checked-in Phase 2 lesson catalog from the planning workbook."""

from __future__ import annotations

import json
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"C:\Users\saisa\Downloads\GeoGebra_Style_Math_App_Pages.xlsx")
OUTPUT = Path(__file__).resolve().parents[1] / "src" / "modules" / "lessons" / "catalog" / "phase2.generated.ts"
PHASE_TWO_CATEGORIES = {
    "Algebra",
    "Graphs and Functions",
    "Geometry",
    "Trigonometry",
    "Symbolic Mathematics",
}

TOPIC_ADAPTERS = {
    "2D Graphing Calculator": "graph",
    "Expressions and Manipulation": "algebra-cas",
    "Equations and Inequalities": "algebra-cas",
    "Functions": "graph",
    "Function Transformations": "graph",
    "Coordinate Geometry": "geometry2d",
    "Vectors": "vector",
    "Dynamic Geometry Constructions": "geometry2d",
    "Transformations and Loci": "geometry2d",
    "Trigonometry": "trigonometry",
    "CAS Workspace": "cas",
}


def slugify(value: str) -> str:
    ascii_value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "-", ascii_value.lower()).strip("-")


def clean_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value)
    if "Ã¢" in text or "Ãƒ" in text:
        try:
            return text.encode("cp1252").decode("utf-8")
        except (UnicodeEncodeError, UnicodeDecodeError):
            pass
    return text


workbook = load_workbook(WORKBOOK, read_only=False, data_only=True)
sheet = workbook["Math Pages Master"]
headers = [cell.value for cell in sheet[1]]
lessons = []

for values in sheet.iter_rows(min_row=2, values_only=True):
    row = dict(zip(headers, values))
    if row["ID"] is None or row["Main Category"] not in PHASE_TWO_CATEGORIES:
        continue
    category = clean_text(row["Main Category"])
    topic = clean_text(row["Topic"])
    title = clean_text(row["Subtopic / Page"])
    assert category and topic and title
    category_slug = slugify(category)
    title_slug = slugify(title)
    lesson_id = int(row["ID"])
    lessons.append({
        "id": lesson_id,
        "phase": 2,
        "slug": title_slug,
        "categorySlug": category_slug,
        "route": f"/lessons/{category_slug}/{lesson_id}-{title_slug}",
        "category": category,
        "topic": topic,
        "title": title,
        "purpose": clean_text(row["Purpose"]),
        "description": clean_text(row["What the Page Does"]),
        "workspace": clean_text(row["Primary Workspace / View"]),
        "interactions": clean_text(row["Core Interactions"]),
        "outcome": clean_text(row["Key Learning Outcome"]),
        "mode": clean_text(row["Suggested Page Mode"]),
        "level": clean_text(row["Level"]),
        "feature": clean_text(row["GeoGebra-Style Feature"]),
        "priority": clean_text(row["Implementation Priority"]),
        "notes": clean_text(row["Notes"]),
        "adapter": TOPIC_ADAPTERS[topic],
    })

lessons.sort(key=lambda lesson: lesson["id"])
ids = [lesson["id"] for lesson in lessons]
assert len(lessons) == 225, f"Expected 225 Phase 2 lessons, got {len(lessons)}"
assert len(ids) == len(set(ids)), "Duplicate Phase 2 IDs"
assert len({lesson["route"] for lesson in lessons}) == len(lessons), "Duplicate Phase 2 routes"

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
payload = json.dumps(lessons, ensure_ascii=False, indent=2)
OUTPUT.write_text(
    "// Generated from GeoGebra_Style_Math_App_Pages.xlsx. Do not edit by hand.\n"
    'import type { LessonSourceDefinition } from "../types";\n\n'
    f"export const phaseTwoLessons = {payload} as const satisfies readonly LessonSourceDefinition[];\n",
    encoding="utf-8",
)
print(f"Generated {len(lessons)} lessons at {OUTPUT}")
