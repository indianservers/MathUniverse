"""Generate the checked-in Phase 1 lesson catalog from the planning workbook."""

from __future__ import annotations

import json
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"C:\Users\saisa\Downloads\GeoGebra_Style_Math_App_Pages.xlsx")
OUTPUT = Path(__file__).resolve().parents[1] / "src" / "modules" / "lessons" / "catalog" / "phase1.generated.ts"
PHASE_ONE_CATEGORIES = {
    "Core Workspaces",
    "Numbers and Arithmetic",
    "Authoring and Learning System",
    "Platform Capabilities",
}


def slugify(value: str) -> str:
    ascii_value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "-", ascii_value.lower()).strip("-")


def clean_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value)
    if "â" in text or "Ã" in text:
        try:
            return text.encode("cp1252").decode("utf-8")
        except (UnicodeEncodeError, UnicodeDecodeError):
            pass
    return text


def adapter_for(topic: str) -> str:
    return {
        "Scientific Calculator": "calculator",
        "Algebra and Dynamic Variables": "algebra",
        "Numbers and Number Theory": "number",
        "Fractions, Decimals, Ratios and Percentages": "number",
        "Interactive Authoring": "authoring",
        "Lesson and Assessment Pages": "learning",
        "Common Tools and Accessibility": "platform",
    }[topic]


workbook = load_workbook(WORKBOOK, read_only=False, data_only=True)
sheet = workbook["Math Pages Master"]
headers = [cell.value for cell in sheet[1]]
lessons = []

for values in sheet.iter_rows(min_row=2, values_only=True):
    row = dict(zip(headers, values))
    if row["ID"] is None or row["Main Category"] not in PHASE_ONE_CATEGORIES:
        continue
    category = clean_text(row["Main Category"])
    topic = clean_text(row["Topic"])
    title = clean_text(row["Subtopic / Page"])
    assert category and topic and title
    category_slug = slugify(category)
    title_slug = slugify(title)
    lesson_id = int(row["ID"])
    lessons.append(
        {
            "id": lesson_id,
            "phase": 1,
            "slug": title_slug,
            "categorySlug": category_slug,
            "route": f"/lessons/{category_slug}/{lesson_id}-{title_slug}",
            "category": category,
            "topic": topic,
            "title": title,
            "purpose": clean_text(row["Purpose"]),
            "description": clean_text(row["What the Page Does"]),
            "workspace": clean_text(row["Primary Workspace / View"]),
            "interactions": clean_text(row["Core Interactions"]),
            "outcome": clean_text(row["Key Learning Outcome"]),
            "mode": clean_text(row["Suggested Page Mode"]),
            "level": clean_text(row["Level"]),
            "feature": clean_text(row["GeoGebra-Style Feature"]),
            "priority": clean_text(row["Implementation Priority"]),
            "notes": clean_text(row["Notes"]),
            "adapter": adapter_for(topic),
        }
    )

lessons.sort(key=lambda lesson: lesson["id"])
ids = [lesson["id"] for lesson in lessons]
assert len(lessons) == 130, f"Expected 130 Phase 1 lessons, got {len(lessons)}"
assert len(ids) == len(set(ids)), "Duplicate Phase 1 IDs"
assert len({lesson["route"] for lesson in lessons}) == len(lessons), "Duplicate Phase 1 routes"

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
payload = json.dumps(lessons, ensure_ascii=False, indent=2)
OUTPUT.write_text(
    "// Generated from GeoGebra_Style_Math_App_Pages.xlsx. Do not edit by hand.\n"
    'import type { LessonSourceDefinition } from "../types";\n\n'
    f"export const phaseOneLessons = {payload} as const satisfies readonly LessonSourceDefinition[];\n",
    encoding="utf-8",
)
print(f"Generated {len(lessons)} lessons at {OUTPUT}")
